"""
Export utilities for generating PDF and Excel reports
"""
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def format_currency(amount):
    """Format amount as Indian currency"""
    return f"₹{amount:,.2f}"

def format_currency_pdf(amount):
    """Format amount for PDF - using Rs. for better compatibility"""
    return f"Rs. {amount:,.2f}"


def generate_outstanding_pdf(results):
    """Generate PDF for Outstanding Report with improved formatting"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           topMargin=0.75*inch, 
                           bottomMargin=0.75*inch,
                           leftMargin=0.5*inch,
                           rightMargin=0.5*inch)
    story = []
    
    # Enhanced Styles
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    # Heading style
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=16,
        fontName='Helvetica-Bold'
    )
    
    # Normal style
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica'
    )
    
    # Currency style (right aligned)
    currency_style = ParagraphStyle(
        'CurrencyStyle',
        parent=normal_style,
        alignment=TA_RIGHT,
        fontSize=10,
        fontName='Helvetica'
    )
    
    # Title Section
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph("OUTSTANDING REPORT", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", subtitle_style))
    story.append(Spacer(1, 0.4*inch))
    
    # Calculate totals
    total_billed = 0
    total_incoming = 0
    total_outgoing = 0
    total_outstanding = 0
    
    # Prepare table data with proper formatting
    table_data = []
    
    # Header row
    header_row = [
        Paragraph('<b>Vendor</b>', normal_style),
        Paragraph('<b>Total Billed</b>', currency_style),
        Paragraph('<b>Total Incoming</b>', currency_style),
        Paragraph('<b>Total Outgoing</b>', currency_style),
        Paragraph('<b>Outstanding</b>', currency_style)
    ]
    table_data.append(header_row)
    
    # Data rows
    for idx, result in enumerate(results):
        vendor_name = result['vendor'].name
        total_billed += result['total_billed']
        total_incoming += result['total_incoming']
        total_outgoing += result['total_outgoing']
        total_outstanding += result['outstanding']
        
        # Color code outstanding amounts
        outstanding_amt = result['outstanding']
        if outstanding_amt > 0:
            outstanding_text = f'<font color="#d32f2f"><b>{format_currency_pdf(outstanding_amt)}</b></font>'
        elif outstanding_amt < 0:
            outstanding_text = f'<font color="#388e3c"><b>{format_currency_pdf(outstanding_amt)}</b></font>'
        else:
            outstanding_text = format_currency_pdf(outstanding_amt)
        
        row = [
            Paragraph(vendor_name, normal_style),
            Paragraph(format_currency_pdf(result['total_billed']), currency_style),
            Paragraph(format_currency_pdf(result['total_incoming']), currency_style),
            Paragraph(format_currency_pdf(result['total_outgoing']), currency_style),
            Paragraph(outstanding_text, currency_style)
        ]
        table_data.append(row)
    
    # Total row
    if total_outstanding > 0:
        total_outstanding_text = f'<font color="#d32f2f"><b>{format_currency_pdf(total_outstanding)}</b></font>'
    elif total_outstanding < 0:
        total_outstanding_text = f'<font color="#388e3c"><b>{format_currency_pdf(total_outstanding)}</b></font>'
    else:
        total_outstanding_text = format_currency_pdf(total_outstanding)
    
    total_row = [
        Paragraph('<b>TOTAL</b>', normal_style),
        Paragraph(f'<b>{format_currency_pdf(total_billed)}</b>', currency_style),
        Paragraph(f'<b>{format_currency_pdf(total_incoming)}</b>', currency_style),
        Paragraph(f'<b>{format_currency_pdf(total_outgoing)}</b>', currency_style),
        Paragraph(f'<b>{total_outstanding_text}</b>', currency_style)
    ]
    table_data.append(total_row)
    
    # Create and style table
    table = Table(table_data, colWidths=[2.8*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
        ('TOPPADDING', (0, 0), (-1, 0), 14),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.white),
        
        # Data rows styling
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('ALIGN', (0, 1), (0, -2), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -2), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('LEFTPADDING', (0, 1), (-1, -2), 10),
        ('RIGHTPADDING', (0, 1), (-1, -2), 10),
        ('TOPPADDING', (0, 1), (-1, -2), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 10),
        
        # Total row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('ALIGN', (0, -1), (0, -1), 'LEFT'),
        ('ALIGN', (1, -1), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
        ('LEFTPADDING', (0, -1), (-1, -1), 10),
        ('RIGHTPADDING', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#34495e')),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.35*inch))
    
    # Summary Section
    story.append(Paragraph("SUMMARY", heading_style))
    story.append(Spacer(1, 0.15*inch))
    
    summary_data = [
        [Paragraph('<b>Total Billed</b>', normal_style), Paragraph(format_currency_pdf(total_billed), currency_style)],
        [Paragraph('<b>Total Incoming Payments</b>', normal_style), Paragraph(format_currency_pdf(total_incoming), currency_style)],
        [Paragraph('<b>Total Outgoing Payments</b>', normal_style), Paragraph(format_currency_pdf(total_outgoing), currency_style)],
        [Paragraph('<b>Net Outstanding</b>', normal_style), 
         Paragraph(f'<b>{total_outstanding_text}</b>', currency_style)]
    ]
    
    summary_table = Table(summary_data, colWidths=[3.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        # Highlight last row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ]))
    story.append(summary_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_collection_pdf(results):
    """Generate PDF for Collection Report with improved formatting"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           topMargin=0.75*inch, 
                           bottomMargin=0.75*inch,
                           leftMargin=0.5*inch,
                           rightMargin=0.5*inch)
    story = []
    
    # Enhanced Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        spaceAfter=4,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica'
    )
    
    currency_style = ParagraphStyle(
        'CurrencyStyle',
        parent=normal_style,
        alignment=TA_RIGHT,
        fontSize=10,
        fontName='Helvetica'
    )
    
    # Title Section
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph("COLLECTION REPORT", title_style))
    story.append(Paragraph(f"Period: {results['start_date'].strftime('%d %B %Y')} to {results['end_date'].strftime('%d %B %Y')}", subtitle_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", date_style))
    story.append(Spacer(1, 0.4*inch))
    
    # Summary table
    summary_data = []
    
    # Header
    summary_data.append([
        Paragraph('<b>Metric</b>', normal_style),
        Paragraph('<b>Amount</b>', currency_style)
    ])
    
    # Data rows with color coding - dark green and dark red for better visibility
    incoming_text = f'<font color="#1b5e20"><b>{format_currency_pdf(results["total_incoming"])}</b></font>'
    outgoing_text = f'<font color="#b71c1c"><b>{format_currency_pdf(results["total_outgoing"])}</b></font>'
    
    net_value = results['net']
    if net_value > 0:
        net_text = f'<font color="#1565c0"><b>{format_currency_pdf(net_value)}</b></font>'
    elif net_value < 0:
        net_text = f'<font color="#d32f2f"><b>{format_currency_pdf(net_value)}</b></font>'
    else:
        net_text = format_currency_pdf(net_value)
    
    summary_data.append([
        Paragraph('Total Incoming', normal_style),
        Paragraph(incoming_text, currency_style)
    ])
    
    summary_data.append([
        Paragraph('Total Outgoing', normal_style),
        Paragraph(outgoing_text, currency_style)
    ])
    
    summary_data.append([
        Paragraph('<b>Net Collection</b>', normal_style),
        Paragraph(f'<b>{net_text}</b>', currency_style)
    ])
    
    summary_table = Table(summary_data, colWidths=[3.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
        ('TOPPADDING', (0, 0), (-1, 0), 14),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.white),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 11),
        ('ALIGN', (0, 1), (0, -2), 'LEFT'),
        ('ALIGN', (1, 1), (1, -2), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.HexColor('#e8f5e9'), colors.HexColor('#ffebee')]),
        ('LEFTPADDING', (0, 1), (-1, -2), 12),
        ('RIGHTPADDING', (0, 1), (-1, -2), 12),
        ('TOPPADDING', (0, 1), (-1, -2), 12),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 12),
        
        # Net row (last row)
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('ALIGN', (0, -1), (0, -1), 'LEFT'),
        ('ALIGN', (1, -1), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, -1), (-1, -1), 14),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 14),
        ('LEFTPADDING', (0, -1), (-1, -1), 12),
        ('RIGHTPADDING', (0, -1), (-1, -1), 12),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#34495e')),
    ]))
    
    story.append(summary_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_deliveries_pdf(stats, delivery_orders=None):
    """Generate PDF for Deliveries Report with improved formatting"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           topMargin=0.75*inch, 
                           bottomMargin=0.75*inch,
                           leftMargin=0.5*inch,
                           rightMargin=0.5*inch)
    story = []
    
    # Enhanced Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=16,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica'
    )
    
    count_style = ParagraphStyle(
        'CountStyle',
        parent=normal_style,
        alignment=TA_CENTER,
        fontSize=11,
        fontName='Helvetica-Bold'
    )
    
    # Title Section
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph("DELIVERIES REPORT", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", date_style))
    story.append(Spacer(1, 0.4*inch))
    
    # Statistics table
    stats_data = []
    
    # Header
    stats_data.append([
        Paragraph('<b>Status</b>', normal_style),
        Paragraph('<b>Count</b>', count_style)
    ])
    
    # Data rows with appropriate colors
    status_data = [
        ('Pending', stats['pending'], '#fff3cd'),
        ('In Transit', stats['in_transit'], '#d1ecf1'),
        ('Delivered', stats['delivered'], '#d4edda'),
        ('Cancelled', stats['cancelled'], '#f8d7da')
    ]
    
    for status, count, bg_color in status_data:
        stats_data.append([
            Paragraph(status, normal_style),
            Paragraph(f'<b>{count}</b>', count_style)
        ])
    
    # Total row
    stats_data.append([
        Paragraph('<b>TOTAL</b>', normal_style),
        Paragraph(f'<b>{stats["total"]}</b>', count_style)
    ])
    
    stats_table = Table(stats_data, colWidths=[3.5*inch, 2.5*inch])
    stats_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
        ('TOPPADDING', (0, 0), (-1, 0), 14),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.white),
        
        # Data rows
        ('FONTNAME', (0, 1), (0, -2), 'Helvetica'),
        ('FONTNAME', (1, 1), (1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 11),
        ('ALIGN', (0, 1), (-1, -2), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [
            colors.HexColor('#fff3cd'),
            colors.HexColor('#d1ecf1'),
            colors.HexColor('#d4edda'),
            colors.HexColor('#f8d7da')
        ]),
        ('LEFTPADDING', (0, 1), (-1, -2), 12),
        ('RIGHTPADDING', (0, 1), (-1, -2), 12),
        ('TOPPADDING', (0, 1), (-1, -2), 12),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 12),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, -1), (-1, -1), 14),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 14),
        ('LEFTPADDING', (0, -1), (-1, -1), 12),
        ('RIGHTPADDING', (0, -1), (-1, -1), 12),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#34495e')),
    ]))
    
    story.append(stats_table)
    
    # Add detailed delivery table if delivery_orders provided
    if delivery_orders:
        story.append(Spacer(1, 0.35*inch))
        story.append(Paragraph("DELIVERY DETAILS", heading_style))
        story.append(Spacer(1, 0.15*inch))
        
        # Detailed table
        details_data = []
        details_data.append([
            Paragraph('<b>Date</b>', normal_style),
            Paragraph('<b>Delivery Person</b>', normal_style),
            Paragraph('<b>Bill/Proxy Bill</b>', normal_style),
            Paragraph('<b>Status</b>', normal_style),
            Paragraph('<b>Address</b>', normal_style)
        ])
        
        for order in delivery_orders:
            # Get delivery person name
            delivery_person = order.delivery_user.username if order.delivery_user else 'N/A'
            
            # Get bill/proxy bill info
            if order.bill:
                bill_info = f"Bill: {order.bill.bill_number}"
            elif order.proxy_bill:
                bill_info = f"Proxy: {order.proxy_bill.proxy_number}"
            else:
                bill_info = "N/A"
            
            # Status color coding
            status_colors = {
                'PENDING': '#f57f17',
                'IN_TRANSIT': '#0277bd',
                'DELIVERED': '#2e7d32',
                'CANCELLED': '#c62828'
            }
            status_color = status_colors.get(order.status, '#000000')
            status_text = f'<font color="{status_color}"><b>{order.status}</b></font>'
            
            # Truncate address if too long
            address = order.delivery_address[:50] + '...' if len(order.delivery_address) > 50 else order.delivery_address
            
            details_data.append([
                Paragraph(order.delivery_date.strftime('%d %b %Y'), normal_style),
                Paragraph(delivery_person, normal_style),
                Paragraph(bill_info, normal_style),
                Paragraph(status_text, normal_style),
                Paragraph(address, normal_style)
            ])
        
        details_table = Table(details_data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1*inch, 2*inch])
        details_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.white),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('LEFTPADDING', (0, 1), (-1, -1), 8),
            ('RIGHTPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(details_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_picklist_pdf(payload):
    """Generate PDF for a single picklist (invoice-style layout)."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=0.75*inch,
                            bottomMargin=0.75*inch,
                            leftMargin=0.5*inch,
                            rightMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    normal_style = ParagraphStyle('NormalStyle', parent=styles['Normal'], fontSize=9, fontName='Helvetica')
    right_style = ParagraphStyle('RightStyle', parent=normal_style, alignment=TA_RIGHT)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceAfter=8, fontName='Helvetica-Bold')

    # Header
    story.append(Paragraph(f"PICKLIST: {payload['picklist_number']}", heading_style))
    story.append(Paragraph(f"Delivery Person: {payload['delivery_person']}  |  Delivery Date: {payload['delivery_date']}", normal_style))
    story.append(Spacer(1, 0.2*inch))

    # Items table
    story.append(Paragraph("Items", heading_style))
    items_data = [['Cat', 'Item Name', 'MRP', 'Value', 'B.UOM', 'CFC', 'PAC', 'FR Qty']]
    for it in payload['items']:
        items_data.append([
            it.get('cat') or '—',
            it.get('item_name') or '—',
            format_currency_pdf(it.get('mrp', 0)),
            format_currency_pdf(it.get('value', 0)),
            it.get('b_uom') or '—',
            it.get('cfc') or '—',
            it.get('pac') or '—',
            str(it.get('fr_qty', '')),
        ])
    items_data.append(['', 'Grand Total', '', format_currency_pdf(payload.get('grand_total', 0)), '', '', '', ''])
    col_widths = [0.4*inch, 2*inch, 0.7*inch, 0.8*inch, 0.5*inch, 0.4*inch, 0.4*inch, 0.5*inch]
    items_table = Table(items_data, colWidths=col_widths)
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 0.25*inch))

    # Invoice details table
    story.append(Paragraph("Invoice Details", heading_style))
    inv_data = [['Invoice No', 'Inv Date', 'Customer Code', 'Customer Name', 'Beat', 'P-Mode', 'InvVal', 'RecAmt']]
    for row in payload.get('invoice_rows', []):
        inv_data.append([
            row.get('invoice_no', ''),
            row.get('inv_date', ''),
            row.get('customer_code', '') or '—',
            row.get('customer_name', '') or '—',
            row.get('beat', '') or '—',
            row.get('p_mode', '') or '—',
            format_currency_pdf(row.get('inv_val', 0)),
            format_currency_pdf(row.get('rec_amt', 0)),
        ])
    inv_col_widths = [0.9*inch, 0.7*inch, 1*inch, 1.5*inch, 0.5*inch, 0.5*inch, 0.7*inch, 0.7*inch]
    inv_table = Table(inv_data, colWidths=inv_col_widths)
    inv_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    story.append(inv_table)
    story.append(Spacer(1, 0.25*inch))

    # Salesman summary
    story.append(Paragraph("Salesman Summary", heading_style))
    ss = payload.get('salesman_summary', {})
    sum_data = [
        ['Salesman Name', 'Invoice Count', 'Net Amount', 'Received Amount'],
        [ss.get('salesman_name', '—'), str(ss.get('invoice_count', 0)),
         format_currency_pdf(ss.get('net_amount', 0)), format_currency_pdf(ss.get('received_amount', 0))],
    ]
    sum_table = Table(sum_data, colWidths=[2*inch, 1*inch, 1.2*inch, 1.2*inch])
    sum_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(sum_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_outstanding_excel(results):
    """Generate Excel for Outstanding Report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Outstanding Report"
    
    # Styles
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Outstanding Report"
    ws['A1'].font = Font(bold=True, size=16, color="2c3e50")
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws['A2'].font = Font(size=10, color="808080")
    ws['A2'].alignment = center_align
    
    # Headers
    headers = ['Vendor', 'Total Billed', 'Total Incoming', 'Total Outgoing', 'Outstanding']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data
    total_billed = 0
    total_incoming = 0
    total_outgoing = 0
    total_outstanding = 0
    
    row_num = 5
    for result in results:
        ws.cell(row=row_num, column=1, value=result['vendor'].name).border = border
        ws.cell(row=row_num, column=2, value=format_currency(result['total_billed'])).alignment = right_align
        ws.cell(row=row_num, column=2).border = border
        ws.cell(row=row_num, column=3, value=format_currency(result['total_incoming'])).alignment = right_align
        ws.cell(row=row_num, column=3).border = border
        ws.cell(row=row_num, column=4, value=format_currency(result['total_outgoing'])).alignment = right_align
        ws.cell(row=row_num, column=4).border = border
        ws.cell(row=row_num, column=5, value=format_currency(result['outstanding'])).alignment = right_align
        ws.cell(row=row_num, column=5).border = border
        
        if result['outstanding'] > 0:
            ws.cell(row=row_num, column=5).font = Font(bold=True, color="DC3545")
        else:
            ws.cell(row=row_num, column=5).font = Font(bold=True, color="28A745")
        
        total_billed += result['total_billed']
        total_incoming += result['total_incoming']
        total_outgoing += result['total_outgoing']
        total_outstanding += result['outstanding']
        row_num += 1
    
    # Total row
    total_row = row_num
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=2, value=format_currency(total_billed)).font = total_font
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).alignment = right_align
    ws.cell(row=total_row, column=2).border = border
    ws.cell(row=total_row, column=3, value=format_currency(total_incoming)).font = total_font
    ws.cell(row=total_row, column=3).fill = total_fill
    ws.cell(row=total_row, column=3).alignment = right_align
    ws.cell(row=total_row, column=3).border = border
    ws.cell(row=total_row, column=4, value=format_currency(total_outgoing)).font = total_font
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).alignment = right_align
    ws.cell(row=total_row, column=4).border = border
    ws.cell(row=total_row, column=5, value=format_currency(total_outstanding)).font = total_font
    ws.cell(row=total_row, column=5).fill = total_fill
    ws.cell(row=total_row, column=5).alignment = right_align
    ws.cell(row=total_row, column=5).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_collection_excel(results):
    """Generate Excel for Collection Report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection Report"
    
    # Styles
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Collection Report"
    ws['A1'].font = Font(bold=True, size=16, color="2c3e50")
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Period: {results['start_date'].strftime('%d %B %Y')} to {results['end_date'].strftime('%d %B %Y')}"
    ws['A2'].font = Font(size=11)
    ws['A2'].alignment = center_align
    
    ws.merge_cells('A3:B3')
    ws['A3'] = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws['A3'].font = Font(size=10, color="808080")
    ws['A3'].alignment = center_align
    
    # Headers
    headers = ['Metric', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data
    data = [
        ['Total Incoming', format_currency(results['total_incoming'])],
        ['Total Outgoing', format_currency(results['total_outgoing'])],
        ['Net Collection', format_currency(results['net'])]
    ]
    
    row_num = 6
    fills = [PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid"),
             PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid"),
             PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid")]
    
    for idx, (metric, amount) in enumerate(data):
        ws.cell(row=row_num, column=1, value=metric).border = border
        ws.cell(row=row_num, column=1).fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=amount).border = border
        ws.cell(row=row_num, column=2).fill = fills[idx]
        ws.cell(row=row_num, column=2).alignment = right_align
        # Dark green for incoming, dark red for outgoing
        if idx == 0:  # Total Incoming
            ws.cell(row=row_num, column=2).font = Font(bold=True, size=11, color="1b5e20")
        elif idx == 1:  # Total Outgoing
            ws.cell(row=row_num, column=2).font = Font(bold=True, size=11, color="b71c1c")
        else:  # Net Collection
            ws.cell(row=row_num, column=2).font = Font(bold=True, size=11)
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_deliveries_excel(stats, delivery_orders=None):
    """Generate Excel for Deliveries Report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliveries Report"
    
    # Styles
    header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Deliveries Report"
    ws['A1'].font = Font(bold=True, size=16, color="2c3e50")
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws['A2'].font = Font(size=10, color="808080")
    ws['A2'].alignment = center_align
    
    # Headers
    headers = ['Status', 'Count']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Data
    data = [
        ('Pending', stats['pending'], 'fff3cd'),
        ('In Transit', stats['in_transit'], 'd1ecf1'),
        ('Delivered', stats['delivered'], 'd4edda'),
        ('Cancelled', stats['cancelled'], 'f8d7da')
    ]
    
    row_num = 5
    for status, count, color in data:
        ws.cell(row=row_num, column=1, value=status).border = border
        ws.cell(row=row_num, column=2, value=count).border = border
        ws.cell(row=row_num, column=2).alignment = center_align
        ws.cell(row=row_num, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=row_num, column=2).font = Font(bold=True, size=11)
        row_num += 1
    
    # Total row
    ws.cell(row=row_num, column=1, value="TOTAL").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=1).border = border
    ws.cell(row=row_num, column=2, value=stats['total']).font = total_font
    ws.cell(row=row_num, column=2).fill = total_fill
    ws.cell(row=row_num, column=2).alignment = center_align
    ws.cell(row=row_num, column=2).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    # Add detailed delivery table if delivery_orders provided
    if delivery_orders:
        # Start details on a new section (after 2 blank rows)
        details_start_row = row_num + 3
        
        # Title for details section
        ws.merge_cells(f'A{details_start_row}:E{details_start_row}')
        ws[f'A{details_start_row}'] = "DELIVERY DETAILS"
        ws[f'A{details_start_row}'].font = Font(bold=True, size=14, color="2c3e50")
        ws[f'A{details_start_row}'].alignment = center_align
        
        # Headers for details table
        details_headers = ['Date', 'Delivery Person', 'Bill/Proxy Bill', 'Status', 'Address']
        details_header_row = details_start_row + 2
        for col_num, header in enumerate(details_headers, 1):
            cell = ws.cell(row=details_header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Data rows
        details_row = details_header_row + 1
        for order in delivery_orders:
            # Get delivery person name
            delivery_person = order.delivery_user.username if order.delivery_user else 'N/A'
            
            # Get bill/proxy bill info
            if order.bill:
                bill_info = f"Bill: {order.bill.bill_number}"
            elif order.proxy_bill:
                bill_info = f"Proxy: {order.proxy_bill.proxy_number}"
            else:
                bill_info = "N/A"
            
            # Status color coding
            status_colors = {
                'PENDING': 'f57f17',
                'IN_TRANSIT': '0277bd',
                'DELIVERED': '2e7d32',
                'CANCELLED': 'c62828'
            }
            status_color = status_colors.get(order.status, '000000')
            
            # Truncate address if too long
            address = order.delivery_address[:60] if len(order.delivery_address) > 60 else order.delivery_address
            
            ws.cell(row=details_row, column=1, value=order.delivery_date.strftime('%d %b %Y')).border = border
            ws.cell(row=details_row, column=2, value=delivery_person).border = border
            ws.cell(row=details_row, column=3, value=bill_info).border = border
            status_cell = ws.cell(row=details_row, column=4, value=order.status)
            status_cell.border = border
            status_cell.font = Font(bold=True, color=status_color)
            ws.cell(row=details_row, column=5, value=address).border = border
            
            # Alternate row colors
            if details_row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=details_row, column=col).fill = PatternFill(
                        start_color="f8f9fa", end_color="f8f9fa", fill_type="solid"
                    )
            
            details_row += 1
        
        # Adjust column widths for details table
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

